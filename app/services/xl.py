import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

def export_to_excel(file, columns, data):
    df = pd.DataFrame(data, columns=columns)

    with pd.ExcelWriter(file, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Données", index=False)
        worksheet = writer.sheets["Données"]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")

        for col_num, col_name in enumerate(columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        for i, col in enumerate(columns, 1):
            max_length = max(df[col].astype(str).map(len).max(), len(col)) + 2
            worksheet.column_dimensions[worksheet.cell(row=1, column=i).column_letter].width = max_length
